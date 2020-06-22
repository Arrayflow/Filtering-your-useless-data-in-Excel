import openpyxl
import numpy as np
from openpyxl import utils


def Openxlsx():
    wb = openpyxl.load_workbook('数据.xlsx')
    ws = wb.active
    return wb, ws

def getData():
    names = locals()
    wb, ws = Openxlsx()
    
    Tx = 0
    allData_2017 = []
    allData_2018 = []
    allData_2019 = []
    
    ##创建97个列表
    for i in range(97):
        names['T2017_'+str(i)] = []
        names['T2018_'+str(i)] = []
        names['T2019_'+str(i)] = []
        
    ##导入每个单元格的数据
    for num in range(2,98):
        cell = utils.get_column_letter(num)
        ##96台获得2017年数据
        for i in range(2,367):
            if ws[cell + str(i)].value is None:
                pass
            else:
                names['T2017_'+str(Tx)].append(ws[cell + str(i)].value)
        ##96台2018年数据
        for i in range(367,732):
            if ws[cell + str(i)].value is None:
                pass
            else:
                names['T2018_'+str(Tx)].append(ws[cell + str(i)].value)
        ##96台2019年数据
        for i in range(732,1077):
            if ws[cell + str(i)].value is None:
                pass
            else:
                names['T2019_'+str(Tx)].append(ws[cell + str(i)].value)
        Tx += 1
    ##最后一台机器单独列出
##    for num in range(2,366):
##        cell = utils.get_column_letter(98)
##        if ws[cell + str(i)].value is None:
##            pass
##        else:
##            T2017_96.append(ws['CT' + str(num)].value)

    for i in range(96):
        allData_2017.append(names['T2017_'+str(i)])
        allData_2018.append(names['T2018_'+str(i)])
        allData_2019.append(names['T2019_'+str(i)])
        
    return allData_2017, allData_2018, allData_2019

def calMean():
    names = locals()
    ##获得3年中所有数据存入列表
    dataT_2017, dataT_2018, dataT_2019 = getData()
    allStd_2017 = []
    allMean_2017 = []
    allStd_2018 = []
    allMean_2018 = []
    allStd_2019 = []
    allMean_2019 = []
    
    ##计算每个单元格的平均值、方差放入大列表
    for i in range(96):
        names['T_2017'+str(i)] = np.array(dataT_2017[i])
        names['T_2018'+str(i)] = np.array(dataT_2018[i])
        names['T_2019'+str(i)] = np.array(dataT_2019[i])
    for i in range(96):
        names['stdT_2017'+str(i)] = np.std(names['T_2017'+str(i)], ddof = 1)
        names['meanT_2017'+str(i)] = sum(dataT_2017[i])/len(dataT_2017[i])
        allStd_2017.append(names['stdT_2017'+str(i)])
        allMean_2017.append(names['meanT_2017'+str(i)])
        
        names['stdT_2018'+str(i)] = np.std(names['T_2018'+str(i)], ddof = 1)
        names['meanT_2018'+str(i)] = sum(dataT_2018[i])/len(dataT_2018[i])
        allStd_2018.append(names['stdT_2018'+str(i)])
        allMean_2018.append(names['meanT_2018'+str(i)])
        
        names['stdT_2019'+str(i)] = np.std(names['T_2019'+str(i)], ddof = 1)
        names['meanT_2019'+str(i)] = sum(dataT_2019[i])/len(dataT_2019[i])
        allStd_2019.append(names['stdT_2019'+str(i)])
        allMean_2019.append(names['meanT_2019'+str(i)])
    ##返回得到17、18、19每一台机器的均值，标准差
    return allStd_2017, allMean_2017, allStd_2018, allMean_2018, allStd_2019, allMean_2019
    
    

def Mainstart():
    
    names = locals()
    
    allStd_2017, allMean_2017, allStd_2018, allMean_2018, allStd_2019, allMean_2019  = calMean()

    allData_2017, allData_2018, allData_2019 = getData()
    
    
    problem_2017 = []
    problem_2018 = []
    problem_2019 = []
    ##创建3年各96个列表装每一列剔除数据
    for i in range(96):
        names['err2017_'+str(i)] = []
        names['err2018_'+str(i)] = []
        names['err2019_'+str(i)] = []
    for num in range(96):
        for k in allData_2017[num]:
            if (k > allMean_2017[num] + allStd_2017[num] * 3) or (k < allMean_2017[num] - allStd_2017[num] * 3):
                names['err2017_'+str(num)].append(k)
            if (k > allMean_2018[num] + allStd_2018[num] * 3) or (k < allMean_2018[num] - allStd_2018[num] * 3):
                names['err2018_'+str(num)].append(k)
            if (k > allMean_2019[num] + allStd_2019[num] * 3) or (k < allMean_2019[num] - allStd_2019[num] * 3):
                names['err2019_'+str(num)].append(k)

    for i in range(96):
        problem_2017.append(names['err2017_'+str(i)])
        problem_2018.append(names['err2018_'+str(i)])
        problem_2019.append(names['err2019_'+str(i)])

    return problem_2017, problem_2018, problem_2019


        

