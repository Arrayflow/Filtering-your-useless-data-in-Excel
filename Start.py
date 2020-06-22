import Data_filter
import openpyxl
from openpyxl import utils
from openpyxl.styles import  PatternFill



def Start():
    
    global kk
    kk = 0
    
    wb, ws = Data_filter.Openxlsx()
    
    problem_2017, problem_2018, problem_2019 = Data_filter.Mainstart()
    
    fill = PatternFill("solid", fgColor="FF0000")

    for i in range(96):
        names['T2017_'+str(i)] = problem_2017[i]
        names['T2018_'+str(i)] = problem_2018[i]
        names['T2019_'+str(i)] = problem_2019[i]
    
    for num in range(2,96):
        
        ##2017年
        for i in range(2,367):
            cell = utils.get_column_letter(num)
            if ws[cell + str(i)].value in names['T2017_'+str(kk)]:
                ws[cell+str(i)].fill = fill
        
            else:
                pass
        
        ##2018年
        for i in range(367,732):
            cell = utils.get_column_letter(num)
            if ws[cell + str(i)].value in names['T2018_'+str(kk)]:
                ws[cell+str(i)].fill = fill
        
            else:
                pass
            
        ##2019年    
        for i in range(732,1077):
            cell = utils.get_column_letter(num)
            if ws[cell + str(i)].value in names['T2019_'+str(kk)]:
                ws[cell+str(i)].fill = fill
        
            else:
                pass
        kk += 1
    wb.save('数据.xlsx')
    
    print(T2017_95)
    print(T2017_94)
    print(T2017_93)
    
    
        
        

if __name__ == '__main__':

    names = locals()

    Start()

